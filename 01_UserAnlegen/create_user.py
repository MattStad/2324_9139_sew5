import random
import string
import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode
"""
@author: Matthias Stadlinger
@klasse: 5CN
"""
def getUsers(file):
    """
    liest Textfile aus
    :param file:
    :return: liste mit Username,Passwort,Student/Lehrer,Klasse
    """
    infos = list()
    wb = load_workbook(file, read_only=True)
    ws = wb[wb.sheetnames[0]]
    first=False
    for row in ws.iter_rows():
        x = row[0]
        y = x.value
        x1 = row[1]
        y1 = x1.value
        x2 = row[2]
        y2 = x2.value
        x3 = row[3]
        y3 = x3.value
        if y != None and first:
            infos.append([y, y1, y2, y3])
        first=True
    return infos


def clean_username(name, existing_usernames):
    """
    reinigt Name
    :param name:
    :param existing_usernames:
    :return: cleaned Name
    """
    # nur Kleinbuchstaben
    cleaned_name = name.lower()
    # keine Sonderzeichen
    cleaned_name = cleaned_name.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
    # Akzente entfernen
    cleaned_name = unidecode(cleaned_name)
    # ' ' mit _ ersetzen
    cleaned_name = cleaned_name.lower().replace(' ', '_')
    # nummer hinzufügen
    suffix = 0
    if cleaned_name in existing_usernames:
        suffix = 1
        while cleaned_name + str(suffix) in existing_usernames:
            suffix += 1
    if suffix != 0:
        cleaned_name += str(suffix)
    return cleaned_name


def generate_password(length=12):
    """
    generiert random passwort
    :param length:
    :return: passwort
    """
    # passwort generieren
    characters = string.ascii_letters + string.digits + string.punctuation
    passwort = '='
    while passwort[0] == '=':
        passwort = ''.join(random.choice(characters) for _ in range(length))
    return passwort

def create_users(student_list, output_type="text"):
    """
    main funktion
    :param student_list:
    :param output_type:
    :return: skript für benutzeranlegen
    """
    """
    >>> test_name = 'Müller'
    >>> cleaned_name = clean_username(test_name, list())
    >>> print(cleaned_name)
    mueller

    >>> test_name = 'ã ñ õ Ã Ñ Õ'
    >>> cleaned_name = clean_username(test_name, list())
    >>> print(cleaned_name)
    a_n_o_a_n_o
    >>> test_name = 'Smith'
    >>> existing = list()
    >>> existing.append("smith")
    >>> existing.append("mueller")
    >>> cleaned_names = clean_username(test_name, existing)
    >>> print(cleaned_names)
    smith1
    >>> test_name = 'Smith'
    >>> existing = list()
    >>> existing.append("smith")
    >>> existing.append("mueller")
    >>> existing.append("smith1")
    >>> cleaned_names = clean_username(test_name, existing)
    >>> print(cleaned_names)
    smith2
    """
    existing_usernames = list()
    users = []
    with open('createUsers', 'w', encoding="utf-8") as f1:
        for student in student_list:
            username = clean_username(student[0], existing_usernames)
            password = generate_password()
            existing_usernames.append(username)

            createAddSkript(f1, student[0], username, password)

            users.append({'Username': username, 'Password': password})

        if output_type == 'text':
            print_users_text(users)
        elif output_type == 'excel':
            create_excel(users)

def createAddSkript(file,username,kusername,password):
    """
    erstellt skript
    :param file:
    :param username:
    :param kusername:
    :param password:
    :return: skript
    """
    file.write("useradd -d \"/home/" + kusername + "\" -c \"" + username + "\" -m -g Klasse -G cdrom,plugdev,sambashare -s /bin/bash " + kusername + "\n")
    file.write("echo \"" + str(password) + ":" + kusername + "\" | chpasswd\n")

def print_users_text(users):
    """
    speichert User und Passwort in Textfile
    :param users:
    :return: Textfile
    """
    with open('Userlist_createuser.txt', 'w') as file:
        for user in users:
            file.write(f"{user['Username']: <25} {user['Password']}\n")

def create_excel(users):
    """
    speicher User und Passwort in Excelfile
    :param users:
    :return: Excelfile
    """
    df = pd.DataFrame(users)
    df.to_excel('Userlist_createuser.xlsx', index=False)


create_users(getUsers("Namen.xlsx"),"text")