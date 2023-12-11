import argparse
import random
from logging.handlers import RotatingFileHandler

from openpyxl import load_workbook
import logging
logger = logging.getLogger('log.log')
logger.setLevel(logging.DEBUG)
logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s')

file_handler = RotatingFileHandler('log.log', maxBytes=10000, backupCount=5)
file_handler.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(console_handler)

def getClasses(file):
    infos=list()
    try:
        wb = load_workbook(file, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows():
            x = row[0]
            y = x.value
            x1 = row[1]
            y1 = x1.value
            x2 = row[2]
            y2 = x2.value

            if y!=None and y[0].isdigit():
                infos.append([y,y1,y2])
        return infos
    except Exception as e:
        logger.error("Fehler beim Lesen des EXELS")
        exit(1)

def createpasswd(klasse,room,teacher):
    zeichenkette = "!%&(),._-=^#"
    return klasse + "" + random.choice(zeichenkette) + "" + str(room) + "" + random.choice(zeichenkette) + "" + teacher + "" + random.choice(zeichenkette)

def createClasses(infos,verbose=False):
    try:
        with open('createClasses', 'w') as f1:
            with open('deleteClasses', 'w') as f2:
                with open('UserList', 'w') as f3:
                    createClassesGrundconfig(f1)
                    deleteClassesGrundconfig(f2)
                    logger.info("Grundconfig erstellt")
                    for i in infos:
                        username=i[0]
                        kusername="k"+username.lower()
                        password=createpasswd(i[0],i[1],i[2])
                        createAddSkript(f1,username,kusername,password)
                        createDelSkript(f2,kusername)
                        createList(f3,kusername, password)
                        if verbose:
                            logger.info("User added to Skripts:"+username)
                    logger.info("UserList erstellt")
                    logger.info("deleteClasses Skript erstellt")
                    logger.info("createClasses Skript erstellt")
    except Exception as e:
        logger.error("Fehler beim erstellen der Klassen")
        exit(1)

def createDelSkript(file,kusername):
    file.write("ueserdel -r "+kusername+"\n")


def createClassesGrundconfig(f):
    f.write("#! /bin/sh\n")
    f.write("groupadd Lehrer\n")
    f.write("groupadd Seminar\n")
    f.write("groupadd Klasse\n")
    f.write("useradd -d \"/home/lehrer\" -c \"lehrer\" -m -g Lehrer -G cdrom,plugdev,sambashare -s /bin/bash lehrer\n")
    f.write("useradd -d \"/home/seminar\" -c \"seminar\" -m -g Seminar -G cdrom,plugdev,sambashare -s /bin/bash seminar\n")

def deleteClassesGrundconfig(f):
    f.write("#! /bin/sh\n")
    f.write("groupdel Lehrer\n")
    f.write("groupdel Seminar\n")
    f.write("groupdel Klasse\n")
    f.write("userdel -r lehrer\n")
    f.write("userdel -r seminar\n")

def createList(file,username,password):
        file.write(username+"      "+password+"\n")

def createAddSkript(file,username,kusername,password):
    file.write("useradd -d \"/home/klassen/" + kusername + "\" -c \"" + username + "\" -m -g Klasse -G cdrom,plugdev,sambashare -s /bin/bash " + kusername + "\n")
    file.write("echo \"" + str(password) + ":" + kusername + "\" | chpasswd\n")


parser = argparse.ArgumentParser(description='Make Classscripts')
parser.add_argument('filename', help='File the classes')
parser.add_argument('-q', '--quite', type=str, help='Print output of every solution')
parser.add_argument('-v', '--verbose', action='store_true',default=False, help='Delay after printing a solution (in milliseconds)')
args = parser.parse_args()

try:
    classes=getClasses(args.filename)
    if(args.quite!=None):
        if(args.quite.upper()=="DEBUG"):
            logger.setLevel(logging.DEBUG)
        elif (args.quite.upper()=="WARNING"):
            logger.setLevel(logging.DEBUG)
        if (args.verbose):
            logger.info("Logging Level set to:"+args.quite)
    createClasses(classes,args.verbose)
except Exception as e:
    print(f"An error occurred: {e}")